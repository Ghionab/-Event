from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.core.mail import send_mail
from django.conf import settings
from django.http import JsonResponse, HttpResponse, Http404
from django.urls import reverse
from django.utils import timezone
from django.db import models
from django.db.models import Q, Count
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt

# Import security decorators
from event_project.decorators import (
    admin_required, 
    organizer_or_admin_required, 
    event_organizer_or_admin_required,
    can_manage_event,
    can_view_registration
)

from .models import (
    Registration, TicketType, PromoCode, RegistrationField, 
    RegistrationDocument, CheckIn, AttendeeNotification, Badge,
    BulkRegistrationUpload, BulkRegistrationRow, ManualRegistration,
    RegistrationStatus
)
from events.models import Event
from .forms import (
    RegistrationForm, TicketTypeForm, PromoCodeForm, 
    RegistrationFieldForm, BulkUploadForm, ManualRegistrationForm
)

# Create your views here.

@can_view_registration
def registration_list(request):
    """List all registrations for the current user"""
    registrations = Registration.objects.filter(
        models.Q(user=request.user) | models.Q(attendee_email=request.user.email)
    ).order_by('-created_at')
    
    return render(request, 'registration/registration_list.html', {'registrations': registrations})

def register_for_event(request, event_id):
    """Register for an event with custom form fields and document uploads"""
    event = get_object_or_404(Event, id=event_id)
    ticket_types = event.ticket_types.filter(is_active=True)

    # Get custom registration fields
    custom_fields = event.registration_fields.filter(is_active=True).order_by('order')

    # Separate file fields from other fields
    file_fields = [f for f in custom_fields if f.field_type == 'file']
    non_file_fields = [f for f in custom_fields if f.field_type != 'file']

    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        form.fields['ticket_type'].queryset = ticket_types

        # Process non-file custom fields
        custom_field_data = {}
        for field in non_file_fields:
            field_key = f"custom_{field.field_name}"
            if field.required and field_key not in request.POST:
                form.add_error(None, f"{field.label} is required")
                continue
            custom_field_data[field.field_name] = request.POST.get(field_key, '')

        # Process file fields - validate but don't save yet
        uploaded_files = {}
        file_errors = {}
        for field in file_fields:
            field_key = f"custom_{field.field_name}"
            uploaded_file = request.FILES.get(field_key) if request.FILES else None

            if field.required and not uploaded_file:
                form.add_error(None, f"{field.label} is required")
                continue

            if uploaded_file:
                # Validate file
                from django.conf import settings
                import os

                # Check file size
                max_size = getattr(settings, 'MAX_UPLOAD_SIZE', 10 * 1024 * 1024)
                if uploaded_file.size > max_size:
                    max_mb = max_size / (1024 * 1024)
                    file_errors[field.label] = f'File size cannot exceed {max_mb}MB'
                    continue

                # Check file extension
                ext = os.path.splitext(uploaded_file.name)[1].lower()
                allowed_exts = field.get_allowed_file_types()
                if ext not in allowed_exts:
                    file_errors[field.label] = f'File type not allowed. Allowed: {", ".join(allowed_exts)}'
                    continue

                uploaded_files[field.field_name] = uploaded_file

        # Add file errors to form
        for error in file_errors.values():
            form.add_error(None, error)

        if form.is_valid() and not file_errors:
            registration = form.save(commit=False)
            registration.event = event
            registration.user = request.user if request.user.is_authenticated else None
            registration.custom_fields = custom_field_data

            # Calculate total amount
            if registration.ticket_type:
                if not registration.ticket_type.can_purchase():
                    messages.error(request, 'Selected ticket type is not available for purchase.')
                    return render(request, 'registration/register_event.html', {
                        'form': form, 'event': event, 'custom_fields': custom_fields,
                        'file_fields': file_fields, 'non_file_fields': non_file_fields
                    })
                total = registration.ticket_type.price
                discount = 0

                # Apply promo code if provided
                promo_code_value = request.POST.get('promo_code', '').strip()
                if promo_code_value:
                    try:
                        promo = PromoCode.objects.get(
                            code=promo_code_value.upper(),
                            event=event,
                            is_active=True
                        )
                        if promo.is_valid():
                            total, discount = promo.apply_discount(total)
                            registration.promo_code = promo
                            promo.current_uses += 1
                            promo.save()
                        else:
                            messages.error(request, 'This promo code is expired or has reached its usage limit.')
                            return render(request, 'registration/register_event.html', {
                                'form': form, 'event': event, 'custom_fields': custom_fields,
                                'file_fields': file_fields, 'non_file_fields': non_file_fields
                            })
                    except PromoCode.DoesNotExist:
                        messages.error(request, 'Invalid promo code.')
                        return render(request, 'registration/register_event.html', {
                            'form': form, 'event': event, 'custom_fields': custom_fields,
                            'file_fields': file_fields, 'non_file_fields': non_file_fields
                        })

                registration.total_amount = total
                registration.discount_amount = discount

            registration.save()
            registration.confirm()

            # Save uploaded files
            for field in file_fields:
                field_key = f"custom_{field.field_name}"
                uploaded_file = uploaded_files.get(field.field_name)
                if uploaded_file:
                    doc = RegistrationDocument(
                        registration=registration,
                        field=field,
                        file=uploaded_file,
                        original_filename=uploaded_file.name,
                        file_name=uploaded_file.name,
                        file_size=uploaded_file.size,
                        mime_type=uploaded_file.content_type
                    )
                    doc.save()
                    # Store document ID in custom_fields for reference
                    custom_field_data[field.field_name] = doc.id

            # Update custom_fields with document references
            registration.custom_fields = custom_field_data
            registration.save()

            messages.success(request, f'Registration successful! Your registration number is {registration.registration_number}')
            return redirect('registration:registration_detail', registration_id=registration.id)
    else:
        form = RegistrationForm()
        form.fields['ticket_type'].queryset = ticket_types

    return render(request, 'registration/register_event.html', {
        'form': form, 'event': event, 'custom_fields': custom_fields,
        'file_fields': file_fields, 'non_file_fields': non_file_fields
    })

def registration_detail(request, registration_id):
    """View registration details"""
    registration = get_object_or_404(Registration, id=registration_id)
    
    # Check permission
    if not request.user.is_authenticated:
        messages.error(request, 'Please log in to view this registration.')
        return redirect('login')
    
    if registration.user and registration.user != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission to view this registration.')
        return redirect('home')
    
    return render(request, 'registration/registration_detail.html', {'registration': registration})

def registration_success(request, registration_id):
    """Display registration success page with QR code"""
    registration = get_object_or_404(Registration, id=registration_id)
    
    # Get event information
    event = registration.event
    
    # Generate QR code image
    qr_code_image = registration.generate_qr_code_image()
    
    # No permission check needed - this is a public success page for the registrant
    
    return render(request, 'participant/registration_success.html', {
        'registration': registration,
        'event': event,
        'qr_code_image': qr_code_image,
    })

def cancel_registration(request, registration_id):
    """Cancel a registration"""
    registration = get_object_or_404(Registration, id=registration_id)
    
    # Check permission
    if not request.user.is_authenticated:
        messages.error(request, 'Please log in to cancel this registration.')
        return redirect('login')
    
    if registration.user and registration.user != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission to cancel this registration.')
        return redirect('registration_detail', registration_id=registration.id)
    
    if request.method == 'POST':
        reason = request.POST.get('reason', '')
        registration.cancel(reason=reason)
        messages.success(request, 'Registration cancelled successfully.')
        return redirect('registration_detail', registration_id=registration.id)
    
    return render(request, 'registration/cancel_registration.html', {'registration': registration})

def ticket_create(request, event_id):
    """Create a ticket type for an event"""
    from events.models import Event
    from organizers.decorators import organizer_required
    
    event = get_object_or_404(Event, id=event_id)
    
    # Check permission
    if not request.user.is_authenticated:
        messages.error(request, 'Please log in to create tickets.')
        return redirect('login')
    
    if event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission to manage tickets for this event.')
        return redirect('event_detail', event_id=event.id)
    
    if request.method == 'POST':
        form = TicketTypeForm(request.POST)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.event = event
            ticket.save()
            messages.success(request, f'Ticket type "{ticket.name}" created successfully!')
            return redirect('organizer_event_detail', event_id=event.id)
    else:
        form = TicketTypeForm()
    
    return render(request, 'registration/ticket_form.html', {'form': form, 'event': event})

def ticket_edit(request, event_id, ticket_id):
    """Edit a ticket type"""
    event = get_object_or_404(Event, id=event_id)
    ticket = get_object_or_404(TicketType, id=ticket_id, event=event)
    
    # Check permission
    if not request.user.is_authenticated:
        messages.error(request, 'Please log in to edit tickets.')
        return redirect('login')
    
    if event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission to manage tickets for this event.')
        return redirect('event_detail', event_id=event.id)
    
    if request.method == 'POST':
        form = TicketTypeForm(request.POST, instance=ticket)
        if form.is_valid():
            form.save()
            messages.success(request, f'Ticket type "{ticket.name}" updated successfully!')
            return redirect('organizer_event_detail', event_id=event.id)
    else:
        form = TicketTypeForm(instance=ticket)
    
    return render(request, 'registration/ticket_form.html', {'form': form, 'event': event, 'ticket': ticket})

def ticket_delete(request, event_id, ticket_id):
    """Delete a ticket type"""
    event = get_object_or_404(Event, id=event_id)
    ticket = get_object_or_404(TicketType, id=ticket_id, event=event)
    
    # Check permission
    if not request.user.is_authenticated:
        messages.error(request, 'Please log in to delete tickets.')
        return redirect('login')
    
    if event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission to manage tickets for this event.')
        return redirect('event_detail', event_id=event.id)
    
    if request.method == 'POST':
        ticket_name = ticket.name
        ticket.delete()
        messages.success(f'Ticket type "{ticket_name}" deleted successfully!')
        return redirect('organizer_event_detail', event_id=event.id)
    
    return render(request, 'registration/ticket_confirm_delete.html', {'ticket': ticket, 'event': event})

def promo_code_create(request, event_id):
    """Create a promo code for an event"""
    event = get_object_or_404(Event, id=event_id)
    
    # Check permission
    if not request.user.is_authenticated:
        messages.error(request, 'Please log in to create promo codes.')
        return redirect('login')
    
    if event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission to manage promo codes for this event.')
        return redirect('event_detail', event_id=event.id)
    
    if request.method == 'POST':
        form = PromoCodeForm(request.POST)
        if form.is_valid():
            promo = form.save(commit=False)
            promo.event = event
            promo.save()
            messages.success(request, f'Promo code "{promo.code}" created successfully!')
            return redirect('organizer_event_detail', event_id=event.id)
    else:
        form = PromoCodeForm()
    
    return render(request, 'registration/promo_code_form.html', {'form': form, 'event': event})

def promo_code_validate(request):
    """Validate a promo code via AJAX"""
    code = request.GET.get('code', '').strip().upper()
    event_id = request.GET.get('event_id')
    ticket_price = request.GET.get('ticket_price', 0)
    
    try:
        promo = PromoCode.objects.get(code=code, event_id=event_id, is_active=True)
        if promo.is_valid():
            discounted_price, discount = promo.apply_discount(float(ticket_price))
            return JsonResponse({
                'valid': True,
                'code': promo.code,
                'discount_type': promo.discount_type,
                'discount_value': str(promo.discount_value),
                'discount_amount': str(discount),
                'discounted_price': str(discounted_price)
            })
        else:
            return JsonResponse({'valid': False, 'message': 'This promo code has expired or reached its usage limit.'})
    except PromoCode.DoesNotExist:
        return JsonResponse({'valid': False, 'message': 'Invalid promo code.'})

def registration_field_create(request, event_id):
    """Create a custom registration field"""
    event = get_object_or_404(Event, id=event_id)
    
    # Check permission
    if not request.user.is_authenticated:
        messages.error(request, 'Please log in to create registration fields.')
        return redirect('login')
    
    if event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission to manage registration fields.')
        return redirect('event_detail', event_id=event.id)
    
    if request.method == 'POST':
        form = RegistrationFieldForm(request.POST)
        if form.is_valid():
            field = form.save(commit=False)
            field.event = event
            field.save()
            messages.success(request, f'Field "{field.label}" created successfully!')
            return redirect('organizer_event_detail', event_id=event.id)
    else:
        form = RegistrationFieldForm()
    
    return render(request, 'registration/registration_field_form.html', {'form': form, 'event': event})

def add_to_waitlist(request, event_id, ticket_type_id):
    """Add user to event waitlist"""
    event = get_object_or_404(Event, id=event_id)
    ticket_type = get_object_or_404(TicketType, id=ticket_type_id, event=event)
    
    if not request.user.is_authenticated:
        messages.error(request, 'Please log in to join the waitlist.')
        return redirect('login')
    
    # Get next position
    last_entry = Waitlist.objects.filter(event=event, ticket_type=ticket_type).order_by('-position').first()
    next_position = (last_entry.position + 1) if last_entry else 1
    
    waitlist_entry, created = Waitlist.objects.get_or_create(
        event=event,
        user=request.user,
        ticket_type=ticket_type,
        defaults={'position': next_position}
    )
    
    if created:
        messages.success(request, f'You have been added to the waitlist! Your position is #{next_position}')
    else:
        messages.info(request, f'You are already on the waitlist at position #{waitlist_entry.position}')
    
    return redirect('event_detail', event_id=event.id)

@login_required
def my_registrations(request):
    """View all registrations for the current user"""
    registrations = Registration.objects.filter(
        models.Q(user=request.user) | models.Q(attendee_email=request.user.email)
    ).order_by('-created_at')
    
    return render(request, 'registration/my_registrations.html', {'registrations': registrations})


# Phase 4 - Attendee Experience Views

@login_required
def attendee_dashboard(request):
    """Attendee dashboard with upcoming events and registrations"""
    user = request.user
    
    # Get user's registrations
    registrations = Registration.objects.filter(
        models.Q(user=user) | models.Q(attendee_email=user.email)
    ).select_related('event', 'ticket_type').order_by('-created_at')
    
    # Separate upcoming and past
    now = timezone.now()
    upcoming = registrations.filter(event__start_date__gte=now, status__in=['confirmed', 'checked_in'])
    past = registrations.filter(event__start_date__lt=now)
    
    # Get saved events (events user is interested in but not registered)
    # This would typically come from a separate model or user profile
    
    context = {
        'upcoming_registrations': upcoming,
        'past_registrations': past,
        'total_registrations': registrations.count(),
    }
    
    return render(request, 'registration/attendee_dashboard.html', context)


@login_required
def attendee_event_detail(request, event_id, registration_id):
    """Detailed view for a specific registration"""
    registration = get_object_or_404(
        Registration.objects.select_related('event', 'ticket_type', 'event__organizer'),
        id=registration_id
    )
    
    # Check permission
    if not request.user.is_authenticated:
        messages.error(request, 'Please log in to view this registration.')
        return redirect('login')
    
    if registration.user and registration.user != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission to view this registration.')
        return redirect('registration:attendee_dashboard')
    
    # Get event sessions
    sessions = registration.event.sessions.filter(is_public=True).order_by('start_time')
    
    # Get speakers
    speakers = registration.event.speakers.filter(is_confirmed=True).order_by('display_order')
    
    # Get badge if exists
    badge = getattr(registration, 'badge', None)
    
    context = {
        'registration': registration,
        'sessions': sessions,
        'speakers': speakers,
        'badge': badge,
    }
    
    return render(request, 'registration/attendee_event_detail.html', context)


@login_required
def badge_view(request, registration_id):
    """View badge for a registration"""
    from .models import Badge
    
    registration = get_object_or_404(
        Registration.objects.select_related('event', 'ticket_type'),
        id=registration_id
    )
    
    # Check permission
    if not request.user.is_authenticated:
        messages.error(request, 'Please log in to view this badge.')
        return redirect('login')
    
    if registration.user and registration.user != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission to view this badge.')
        return redirect('attendee_dashboard')
    
    # Get or create badge
    badge, created = Badge.objects.get_or_create(
        registration=registration,
        defaults={
            'name': registration.attendee_name,
            'badge_type': 'vip' if registration.ticket_type and registration.ticket_type.ticket_category == 'vip' else 'standard',
            'qr_code_data': f"REG:{registration.qr_code}",
        }
    )
    
    # Generate QR code
    qr_code_image = badge.generate_qr_code()
    
    context = {
        'badge': badge,
        'registration': registration,
        'qr_code_image': qr_code_image,
    }
    
    return render(request, 'registration/badge_view.html', context)


@login_required
def badge_print(request, registration_id):
    """Print badge for a registration"""
    from .models import Badge
    
    registration = get_object_or_404(Registration, id=registration_id)
    
    # Check permission
    if not request.user.is_authenticated:
        messages.error(request, 'Please log in to print this badge.')
        return redirect('login')
    
    if registration.user and registration.user != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission to print this badge.')
        return redirect('attendee_dashboard')
    
    # Get or create badge
    badge, created = Badge.objects.get_or_create(
        registration=registration,
        defaults={
            'name': registration.attendee_name,
            'badge_type': 'vip' if registration.ticket_type and registration.ticket_type.ticket_category == 'vip' else 'standard',
            'qr_code_data': f"REG:{registration.qr_code}",
        }
    )
    
    # Mark as printed
    badge.mark_printed(request.user)
    
    messages.success(request, 'Badge marked as printed.')
    return redirect('registration:badge_view', registration_id=registration.id)


def qr_check_in(request):
    """Handle QR code check-in (for kiosks or scanning devices)"""
    qr_data = request.GET.get('qr', '')
    
    if not qr_data:
        return JsonResponse({'success': False, 'message': 'No QR code data provided'})
    
    # Try to find registration by QR code
    try:
        registration = Registration.objects.select_related('event', 'ticket_type').get(qr_code=qr_data)
        
        if registration.status == 'checked_in':
            return JsonResponse({
                'success': False,
                'message': 'Already checked in',
                'registration': {
                    'name': registration.attendee_name,
                    'event': registration.event.title,
                    'checked_in_at': registration.checked_in_at.isoformat(),
                }
            })
        
        # Perform check-in
        registration.check_in(checked_by=None)
        
        # Log check-in
        CheckIn.objects.create(
            registration=registration,
            checked_in_by=None,
            method='qr_scan'
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Check-in successful!',
            'registration': {
                'name': registration.attendee_name,
                'event': registration.event.title,
                'ticket_type': registration.ticket_type.name if registration.ticket_type else 'N/A',
            }
        })
        
    except Registration.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Invalid QR code'})


@login_required
def manual_check_in(request, event_id):
    """Manual check-in page for organizers"""
    from events.models import Event
    from .models import CheckIn
    
    event = get_object_or_404(Event, id=event_id)
    
    # Check permission
    if not request.user.is_authenticated:
        messages.error(request, 'Please log in to check in attendees.')
        return redirect('login')
    
    if event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission to check in attendees for this event.')
        return redirect('event_detail', event_id=event.id)
    
    search_query = request.GET.get('q', '')
    checkin_results = None
    
    if search_query:
        # Search by registration number, email, or name
        registrations = Registration.objects.filter(
            event=event,
            status='confirmed'
        ).filter(
            models.Q(registration_number__icontains=search_query) |
            models.Q(attendee_email__icontains=search_query) |
            models.Q(attendee_name__icontains=search_query)
        )[:20]
        
        checkin_results = registrations
    
    # Get check-in stats
    total_confirmed = Registration.objects.filter(event=event, status='confirmed').count()
    total_checked_in = Registration.objects.filter(event=event, status='checked_in').count()
    
    context = {
        'event': event,
        'search_query': search_query,
        'checkin_results': checkin_results,
        'total_confirmed': total_confirmed,
        'total_checked_in': total_checked_in,
        'check_in_rate': (total_checked_in / total_confirmed * 100) if total_confirmed > 0 else 0,
    }
    
    return render(request, 'registration/manual_checkin.html', context)


@login_required
def perform_check_in(request, registration_id):
    """Perform check-in for a specific registration"""
    registration = get_object_or_404(Registration, id=registration_id)
    
    # Check permission
    if not request.user.is_authenticated:
        messages.error(request, 'Please log in to check in attendees.')
        return redirect('login')
    
    if registration.event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission to check in attendees.')
        return redirect('manual_check_in', event_id=registration.event.id)
    
    if registration.status == 'checked_in':
        messages.warning(request, f'{registration.attendee_name} is already checked in.')
    elif registration.check_in(checked_by=request.user):
        # Log check-in
        CheckIn.objects.create(
            registration=registration,
            checked_in_by=request.user,
            method='manual'
        )
        messages.success(f'{registration.attendee_name} checked in successfully!')
    else:
        messages.error(f'Cannot check in {registration.attendee_name}. Registration is not confirmed.')
    
    return redirect('manual_check_in', event_id=registration.event.id)


@login_required
def attendee_preferences(request, event_id):
    """Manage attendee preferences for an event"""
    from .models import AttendeePreference
    from events.models import Event
    
    event = get_object_or_404(Event, id=event_id)
    
    # Get or create preferences
    preferences, created = AttendeePreference.objects.get_or_create(
        user=request.user,
        event=event,
        defaults={}
    )
    
    if request.method == 'POST':
        # Update preferences
        preferences.interested_topics = request.POST.getlist('interested_topics', [])
        preferences.preferred_tracks = request.POST.getlist('preferred_tracks', [])
        preferences.dietary_requirements = request.POST.getlist('dietary_requirements', [])
        preferences.dietary_notes = request.POST.get('dietary_notes', '')
        preferences.accessibility_needs = request.POST.getlist('accessibility_needs', [])
        preferences.accessibility_notes = request.POST.get('accessibility_notes', '')
        preferences.networking_enabled = 'networking_enabled' in request.POST
        preferences.networking_bio = request.POST.get('networking_bio', '')
        preferences.linkedin_url = request.POST.get('linkedin_url', '')
        preferences.twitter_handle = request.POST.get('twitter_handle', '')
        preferences.email_notifications = 'email_notifications' in request.POST
        preferences.sms_notifications = 'sms_notifications' in request.POST
        preferences.save()
        
        messages.success(request, 'Preferences saved successfully!')
        return redirect('registration:attendee_event_detail', event_id=event.id, registration_id=0)
    
    # Get event tracks for preference selection
    tracks = event.tracks.all()
    
    context = {
        'event': event,
        'preferences': preferences,
        'tracks': tracks,
    }
    
    return render(request, 'registration/attendee_preferences.html', context)


@login_required
def attendee_messages(request):
    """View and send messages to other attendees"""
    from .models import AttendeeMessage
    
    user = request.user
    
    # Get messages
    received_messages = AttendeeMessage.objects.filter(
        recipient=user
    ).select_related('sender', 'event').order_by('-created_at')
    
    unread_count = received_messages.filter(is_read=False).count()
    
    context = {
        'messages': received_messages,
        'unread_count': unread_count,
    }
    
    return render(request, 'registration/attendee_messages.html', context)


@login_required
def send_message(request, recipient_id):
    """Send a message to another attendee"""
    from .models import AttendeeMessage
    from users.models import User
    
    recipient = get_object_or_404(User, id=recipient_id)
    
    if request.method == 'POST':
        subject = request.POST.get('subject', '')
        message = request.POST.get('message', '')
        event_id = request.POST.get('event_id')
        
        event = None
        if event_id:
            event = get_object_or_404(Event, id=event_id)
        
        AttendeeMessage.objects.create(
            sender=request.user,
            recipient=recipient,
            event=event,
            subject=subject,
            message=message
        )
        
        messages.success(request, f'Message sent to {recipient.email}')
        return redirect('registration:attendee_messages')
    
    context = {
        'recipient': recipient,
    }
    
    return render(request, 'registration/send_message.html', context)


@login_required
def mark_message_read(request, message_id):
    """Mark a message as read"""
    from .models import AttendeeMessage
    
    message = get_object_or_404(AttendeeMessage, id=message_id, recipient=request.user)
    
    if not message.is_read:
        message.is_read = True
        message.read_at = timezone.now()
        message.save()

    return redirect('registration:attendee_messages')


@login_required
def session_feedback(request, session_id):
    """Submit feedback for a session"""
    from events.models import EventSession
    from .models import SessionAttendance
    
    session = get_object_or_404(EventSession, id=session_id)
    
    # Get user's registration for this event
    registration = get_object_or_404(
        Registration,
        event=session.event,
        user=request.user,
        status__in=['confirmed', 'checked_in']
    )
    
    # Get or create attendance record
    attendance, created = SessionAttendance.objects.get_or_create(
        registration=registration,
        session=session
    )
    
    if request.method == 'POST':
        rating = request.POST.get('rating')
        feedback = request.POST.get('feedback', '')
        
        if rating:
            attendance.rating = int(rating)
            attendance.feedback = feedback
            attendance.save()
            messages.success(request, 'Thank you for your feedback!')
            return redirect('registration:attendee_event_detail', event_id=session.event.id, registration_id=registration.id)
    
    context = {
        'session': session,
        'attendance': attendance,
    }
    
    return render(request, 'registration/session_feedback.html', context)


# Import CheckIn model at the end to avoid circular imports
from .models import Badge, CheckIn, AttendeePreference, AttendeeMessage, SessionAttendance


# =============================================================================
# Document Upload Views
# =============================================================================

@login_required
def document_list(request, registration_id):
    """List all documents for a registration"""
    registration = get_object_or_404(Registration, id=registration_id)

    # Check permission (owner or organizer)
    is_organizer = False
    if registration.event.organizer == request.user:
        is_organizer = True
    elif registration.user and registration.user != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission to view these documents.')
        return redirect('home')

    documents = registration.documents.all()

    return render(request, 'registration/document_list.html', {
        'registration': registration,
        'documents': documents,
        'is_organizer': is_organizer,
    })


@login_required
def document_detail(request, document_id):
    """View document details"""
    document = get_object_or_404(RegistrationDocument, id=document_id)
    registration = document.registration

    # Check permission
    is_organizer = False
    if registration.event.organizer == request.user:
        is_organizer = True
    elif registration.user and registration.user != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission to view this document.')
        return redirect('home')

    return render(request, 'registration/document_detail.html', {
        'document': document,
        'registration': registration,
        'is_organizer': is_organizer,
    })


@login_required
def document_validate(request, document_id):
    """Validate a document (organizer only)"""
    document = get_object_or_404(RegistrationDocument, id=document_id)
    registration = document.registration

    # Check if user is organizer
    if registration.event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission to validate documents.')
        return redirect('home')

    if request.method == 'POST':
        validation_notes = request.POST.get('validation_notes', '')
        action = request.POST.get('action', 'approve')

        document.is_validated = (action == 'approve')
        document.validation_notes = validation_notes
        document.validated_by = request.user
        document.validated_at = timezone.now()
        document.save()

        messages.success(request, f'Document {"approved" if document.is_validated else "rejected"} successfully.')
        return redirect('registration:document_list', registration_id=registration.id)

    return render(request, 'registration/document_validate.html', {
        'document': document,
        'registration': registration,
    })


@login_required
def document_download(request, document_id):
    """Download a document"""
    document = get_object_or_404(RegistrationDocument, id=document_id)
    registration = document.registration

    # Check permission
    if registration.event.organizer != request.user and \
       not (registration.user and registration.user == request.user) and \
       not request.user.is_staff:
        messages.error(request, 'You do not have permission to download this document.')
        return redirect('home')

    from django.http import FileResponse
    import os

    # Check if file exists
    if not document.file or not os.path.exists(document.file.path):
        messages.error(request, 'File not found.')
        return redirect('registration:document_detail', document_id=document.id)

    # Open file for download
    file_handle = open(document.file.path, 'rb')
    response = FileResponse(file_handle, content_type=document.mime_type or 'application/octet-stream')
    response['Content-Disposition'] = f'attachment; filename="{document.original_filename}"'
    response['Content-Length'] = document.file_size

    return response


@login_required
def document_delete(request, document_id):
    """Delete a document (owner only, before registration is confirmed)"""
    document = get_object_or_404(RegistrationDocument, id=document_id)
    registration = document.registration

    # Only the registration owner can delete their documents
    if registration.user != request.user:
        messages.error(request, 'You do not have permission to delete this document.')
        return redirect('home')

    # Check if registration is still pending
    if registration.status != RegistrationStatus.PENDING:
        messages.error(request, 'Cannot delete documents for confirmed registrations.')
        return redirect('registration:document_list', registration_id=registration.id)

    if request.method == 'POST':
        # Delete the file from storage
        if document.file:
            document.file.delete()
        document.delete()
        messages.success(request, 'Document deleted successfully.')
        return redirect('registration:document_list', registration_id=registration.id)

    return render(request, 'registration/document_delete.html', {
        'document': document,
        'registration': registration,
    })


@login_required
def registration_documents_api(request, registration_id):
    """API endpoint for document operations"""
    registration = get_object_or_404(Registration, id=registration_id)

    # Check permission
    if registration.event.organizer != request.user and \
       not (registration.user and registration.user == request.user) and \
       not request.user.is_staff:
        return JsonResponse({'error': 'Permission denied'}, status=403)

    if request.method == 'GET':
        # List documents
        documents = registration.documents.all()
        data = [{
            'id': doc.id,
            'field_name': doc.field.label,
            'filename': doc.original_filename,
            'file_size': doc.file_size,
            'mime_type': doc.mime_type,
            'is_validated': doc.is_validated,
            'validation_notes': doc.validation_notes,
            'uploaded_at': doc.uploaded_at.isoformat(),
            'validated_at': doc.validated_at.isoformat() if doc.validated_at else None,
        } for doc in documents]
        return JsonResponse({'documents': data})

    elif request.method == 'POST':
        # Upload new document
        field_id = request.POST.get('field_id')
        if not field_id:
            return JsonResponse({'error': 'Field ID required'}, status=400)

        try:
            field = RegistrationField.objects.get(id=field_id, event=registration.event)
        except RegistrationField.DoesNotExist:
            return JsonResponse({'error': 'Invalid field'}, status=400)

        if field.field_type != 'file':
            return JsonResponse({'error': 'This field is not a file upload field'}, status=400)

        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return JsonResponse({'error': 'No file uploaded'}, status=400)

        # Validate file
        from django.conf import settings
        import os

        max_size = getattr(settings, 'MAX_UPLOAD_SIZE', 10 * 1024 * 1024)
        if uploaded_file.size > max_size:
            return JsonResponse({'error': 'File too large'}, status=400)

        ext = os.path.splitext(uploaded_file.name)[1].lower()
        allowed = field.get_allowed_file_types()
        if ext not in allowed:
            return JsonResponse({'error': 'File type not allowed'}, status=400)

        # Create document
        document = RegistrationDocument(
            registration=registration,
            field=field,
            file=uploaded_file,
            original_filename=uploaded_file.name,
            file_name=uploaded_file.name,
            file_size=uploaded_file.size,
            mime_type=uploaded_file.content_type
        )
        document.save()

        return JsonResponse({
            'success': True,
            'document': {
                'id': document.id,
                'filename': document.original_filename,
                'file_size': document.file_size,
            }
        })

    elif request.method == 'DELETE':
        # Delete document
        document_id = request.GET.get('document_id')
        try:
            document = RegistrationDocument.objects.get(id=document_id, registration=registration)
        except RegistrationDocument.DoesNotExist:
            return JsonResponse({'error': 'Document not found'}, status=404)

        # Only owner can delete
        if registration.user != request.user:
            return JsonResponse({'error': 'Permission denied'}, status=403)

        if document.file:
            document.file.delete()
        document.delete()

        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Method not allowed'}, status=405)


# =============================================================================
# Bulk Registration Views
# =============================================================================

@login_required
def bulk_registration_upload(request, event_id):
    """Upload bulk registrations from Excel/CSV file"""
    event = get_object_or_404(Event, id=event_id)

    # Check permission (organizer only)
    if event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission to manage this event.')
        return redirect('home')

    from .forms import BulkUploadForm

    if request.method == 'POST':
        form = BulkUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            file_type = form.cleaned_data['file_type']
            skip_header = form.cleaned_data['skip_header']
            send_invitations = form.cleaned_data['send_invitations']

            # Create bulk upload record
            bulk_upload = BulkRegistrationUpload.objects.create(
                event=event,
                uploaded_by=request.user,
                file=file,
                original_filename=file.name,
                file_name=file.name,
                file_size=file.size,
                status='pending',
            )

            # Process the file in background (for now, process immediately)
            try:
                process_bulk_registration(bulk_upload, skip_header, send_invitations)
                messages.success(request, f'Bulk registration completed! {bulk_upload.success_count} registered, {bulk_upload.error_count} errors.')
            except Exception as e:
                bulk_upload.status = 'failed'
                bulk_upload.error_log = str(e)
                bulk_upload.save()
                messages.error(request, f'Error processing file: {e}')

            return redirect('registration:bulk_registration_detail', bulk_id=bulk_upload.id)
    else:
        form = BulkUploadForm()

    # Get upload history
    uploads = BulkRegistrationUpload.objects.filter(event=event).order_by('-created_at')[:10]

    return render(request, 'registration/bulk_upload.html', {
        'event': event,
        'form': form,
        'uploads': uploads,
    })


def process_bulk_registration(bulk_upload, skip_header=True, send_invitations=True):
    """Process bulk registration from uploaded file"""
    import os
    from openpyxl import load_workbook
    import csv

    bulk_upload.status = 'processing'
    bulk_upload.save()

    file_path = bulk_upload.file.path
    file_ext = os.path.splitext(bulk_upload.original_filename)[1].lower()

    rows_data = []
    header_row = None

    try:
        # Read file based on type
        if file_ext in ['.xlsx', '.xls']:
            wb = load_workbook(filename=file_path, data_only=True)
            ws = wb.active

            # Get header row
            if skip_header:
                header_row = [str(cell.value).strip().lower() if cell.value else f'col_{i}' for i, cell in enumerate(ws[1])]
                start_row = 2
            else:
                header_row = ['name', 'email', 'phone', 'company', 'job_title']
                start_row = 1

            # Read data rows
            for row in ws.iter_rows(min_row=start_row):
                row_data = {}
                for i, cell in enumerate(row):
                    if i < len(header_row):
                        row_data[header_row[i]] = str(cell.value).strip() if cell.value is not None else ''
                if any(row_data.values()):
                    rows_data.append(row_data)

        elif file_ext == '.csv':
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)

                if skip_header:
                    header_row = [str(h).strip().lower() if h else f'col_{i}' for i, h in enumerate(next(reader))]
                else:
                    header_row = ['name', 'email', 'phone', 'company', 'job_title']

                for row in reader:
                    row_data = {}
                    for i, cell in enumerate(row):
                        if i < len(header_row):
                            row_data[header_row[i]] = cell.strip()
                    if any(row_data.values()):
                        rows_data.append(row_data)

        # Create row records
        bulk_upload.total_rows = len(rows_data)
        bulk_upload.save()

        success_count = 0
        error_count = 0
        error_log = []

        # Process each row
        for i, row_data in enumerate(rows_data, 1):
            bulk_row = BulkRegistrationRow.objects.create(
                bulk_upload=bulk_upload,
                row_number=i,
                row_data=row_data,
                status='pending'
            )

            try:
                # Extract fields
                name = row_data.get('name', row_data.get('full_name', ''))
                email = row_data.get('email', row_data.get('email_address', ''))
                phone = row_data.get('phone', row_data.get('telephone', ''))
                company = row_data.get('company', row_data.get('organization', ''))
                job_title = row_data.get('job_title', row_data.get('title', ''))

                if not name or not email:
                    raise ValueError('Name and email are required')

                # Check if already registered
                existing = Registration.objects.filter(
                    event=bulk_upload.event,
                    attendee_email__iexact=email
                ).first()

                if existing:
                    raise ValueError(f'Email {email} already registered')

                # Get default ticket type
                ticket_type = bulk_upload.event.ticket_types.filter(
                    is_active=True
                ).first()

                if not ticket_type:
                    raise ValueError('No active ticket type available')

                # Create registration
                registration = Registration.objects.create(
                    event=bulk_upload.event,
                    attendee_name=name,
                    attendee_email=email,
                    attendee_phone=phone or '',
                    custom_fields={
                        'company': company or '',
                        'job_title': job_title or '',
                    },
                    ticket_type=ticket_type,
                    total_amount=ticket_type.price,
                    status=RegistrationStatus.CONFIRMED
                )

                # Update ticket count
                ticket_type.quantity_sold += 1
                ticket_type.save()

                # Send invitation if requested
                if send_invitations:
                    try:
                        from django.core.mail import send_mail
                        from django.conf import settings

                        subject = f"Event Registration Confirmed: {bulk_upload.event.title}"
                        message = f"""
Dear {name},

Your registration for {bulk_upload.event.title} has been confirmed!

Event Details:
- Date: {bulk_upload.event.start_date.strftime('%B %d, %Y')}
- Venue: {bulk_upload.event.venue_name or 'TBA'}
- Ticket: {ticket_type.name}

Registration Number: {registration.registration_number}

Best regards,
{bulk_upload.event.organizer.get_full_name() or 'Event Team'}
                        """
                        send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [email], fail_silently=True)
                    except Exception:
                        pass  # Don't fail if email fails

                # Link to row
                bulk_row.registration = registration
                bulk_row.status = 'success'
                bulk_row.processed_at = timezone.now()
                bulk_row.save()

                success_count += 1

            except Exception as e:
                bulk_row.status = 'failed'
                bulk_row.error_message = str(e)
                bulk_row.processed_at = timezone.now()
                bulk_row.save()

                error_count += 1
                error_log.append(f"Row {i}: {str(e)}")

        bulk_upload.status = 'completed'
        bulk_upload.success_count = success_count
        bulk_upload.error_count = error_count
        bulk_upload.error_log = '\n'.join(error_log)
        bulk_upload.completed_at = timezone.now()
        bulk_upload.save()

    except Exception as e:
        bulk_upload.status = 'failed'
        bulk_upload.error_log = str(e)
        bulk_upload.save()
        raise


@login_required
def bulk_registration_list(request, event_id):
    """List all bulk uploads for an event"""
    event = get_object_or_404(Event, id=event_id)

    # Check permission
    if event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission.')
        return redirect('home')

    uploads = BulkRegistrationUpload.objects.filter(event=event)

    return render(request, 'registration/bulk_upload_list.html', {
        'event': event,
        'uploads': uploads,
    })


@login_required
def bulk_registration_detail(request, bulk_id):
    """View details of a bulk upload"""
    bulk_upload = get_object_or_404(BulkRegistrationUpload, id=bulk_id)

    # Check permission
    if bulk_upload.event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission.')
        return redirect('home')

    rows = bulk_upload.rows.all()

    return render(request, 'registration/bulk_upload_detail.html', {
        'bulk_upload': bulk_upload,
        'rows': rows,
    })


@login_required
def manual_registration_create(request, event_id):
    """Create a manual registration"""
    event = get_object_or_404(Event, id=event_id)

    # Check permission
    if event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission.')
        return redirect('home')

    from .forms import ManualRegistrationForm

    if request.method == 'POST':
        form = ManualRegistrationForm(request.POST, event=event)
        if form.is_valid():
            manual_reg = form.save(commit=False)
            manual_reg.event = event
            manual_reg.created_by = request.user

            # Check if should create registration immediately
            if request.POST.get('create_registration'):
                registration = manual_reg.create_registration()
                if registration:
                    messages.success(request, f'Registration created! Number: {registration.registration_number}')
                else:
                    messages.warning(request, 'Manual entry saved but no ticket type selected')
            else:
                manual_reg.save()
                if request.POST.get('send_invite'):
                    if manual_reg.send_invite():
                        messages.success(request, f'Invitation sent to {manual_reg.attendee_email}')
                    else:
                        messages.error(request, 'Failed to send invitation email')

                messages.success(request, 'Manual registration entry saved')

            return redirect('registration:manual_registration_list', event_id=event.id)
    else:
        form = ManualRegistrationForm(event=event)

    return render(request, 'registration/manual_registration_form.html', {
        'event': event,
        'form': form,
    })


@login_required
def manual_registration_list(request, event_id):
    """List all manual registrations for an event"""
    event = get_object_or_404(Event, id=event_id)

    # Check permission
    if event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission.')
        return redirect('home')

    manual_regs = ManualRegistration.objects.filter(event=event)

    return render(request, 'registration/manual_registration_list.html', {
        'event': event,
        'manual_regs': manual_regs,
    })


@login_required
def manual_registration_edit(request, event_id, reg_id):
    """Edit a manual registration"""
    event = get_object_or_404(Event, id=event_id)
    manual_reg = get_object_or_404(ManualRegistration, id=reg_id, event=event)

    # Check permission
    if event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission.')
        return redirect('home')

    from .forms import ManualRegistrationForm

    if request.method == 'POST':
        form = ManualRegistrationForm(request.POST, instance=manual_reg, event=event)
        if form.save():
            messages.success(request, 'Manual registration updated')
            return redirect('registration:manual_registration_list', event_id=event.id)
    else:
        form = ManualRegistrationForm(instance=manual_reg, event=event)

    return render(request, 'registration/manual_registration_form.html', {
        'event': event,
        'form': form,
        'manual_reg': manual_reg,
    })


@login_required
def manual_registration_delete(request, event_id, reg_id):
    """Delete a manual registration"""
    event = get_object_or_404(Event, id=event_id)
    manual_reg = get_object_or_404(ManualRegistration, id=reg_id, event=event)

    # Check permission
    if event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission.')
        return redirect('home')

    if request.method == 'POST':
        # Delete associated registration if exists
        if manual_reg.registration:
            manual_reg.registration.delete()

        manual_reg.delete()
        messages.success(request, 'Manual registration deleted')
        return redirect('registration:manual_registration_list', event_id=event.id)

    return render(request, 'registration/manual_registration_confirm_delete.html', {
        'event': event,
        'manual_reg': manual_reg,
    })


@login_required
def manual_registration_send_invite(request, event_id, reg_id):
    """Send invitation for manual registration"""
    event = get_object_or_404(Event, id=event_id)
    manual_reg = get_object_or_404(ManualRegistration, id=reg_id, event=event)

    # Check permission
    if event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission.')
        return redirect('home')

    if manual_reg.send_invite():
        messages.success(request, f'Invitation sent to {manual_reg.attendee_email}')
    else:
        messages.error(request, 'Failed to send invitation email')

    return redirect('registration:manual_registration_list', event_id=event.id)


@login_required
def manual_registration_create_registration(request, event_id, reg_id):
    """Create actual registration from manual entry"""
    event = get_object_or_404(Event, id=event_id)
    manual_reg = get_object_or_404(ManualRegistration, id=reg_id, event=event)

    # Check permission
    if event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission.')
        return redirect('home')

    if manual_reg.registration:
        messages.warning(request, 'Registration already exists')
    else:
        registration = manual_reg.create_registration()
        if registration:
            messages.success(request, f'Registration created! Number: {registration.registration_number}')
        else:
            messages.error(request, 'Failed to create registration')

    return redirect('registration:manual_registration_list', event_id=event.id)


# Import new models at end
from .models import BulkRegistrationUpload, BulkRegistrationRow, ManualRegistration


# =============================================================================
# QR Code Management Views
# =============================================================================

@login_required
def qr_code_list(request, event_id):
    """List all QR codes for an event"""
    event = get_object_or_404(Event, id=event_id)

    # Check permission
    if event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission.')
        return redirect('home')

    # Get all registrations
    registrations = Registration.objects.filter(event=event).select_related('ticket_type')

    # Filter by status
    status_filter = request.GET.get('status')
    if status_filter:
        registrations = registrations.filter(status=status_filter)

    # Search
    search = request.GET.get('search')
    if search:
        registrations = registrations.filter(
            models.Q(attendee_name__icontains=search) |
            models.Q(attendee_email__icontains=search) |
            models.Q(registration_number__icontains=search)
        )

    # Stats
    total = registrations.count()
    checked_in = registrations.filter(status='checked_in').count()
    pending = registrations.filter(status='pending').count()
    confirmed = registrations.filter(status='confirmed').count()

    return render(request, 'registration/qr_code_list.html', {
        'event': event,
        'registrations': registrations,
        'total': total,
        'checked_in': checked_in,
        'pending': pending,
        'confirmed': confirmed,
        'status_filter': status_filter,
        'search': search,
    })


@login_required
def qr_code_download(request, registration_id):
    """Download QR code image for a registration"""
    registration = get_object_or_404(Registration, id=registration_id)

    # Check permission
    if registration.event.organizer != request.user and \
       not (registration.user == request.user) and \
       not request.user.is_staff:
        messages.error(request, 'You do not have permission.')
        return redirect('home')

    # Generate QR code
    qr_image = registration.generate_qr_code_image()

    return render(request, 'registration/qr_code_download.html', {
        'registration': registration,
        'qr_image': qr_image,
    })


@login_required
def qr_code_print(request, event_id):
    """Print QR codes for all registrations"""
    event = get_object_or_404(Event, id=event_id)

    # Check permission
    if event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission.')
        return redirect('home')

    # Get registrations to print
    registrations = Registration.objects.filter(
        event=event,
        status__in=['pending', 'confirmed']
    ).select_related('ticket_type')

    return render(request, 'registration/qr_code_print.html', {
        'event': event,
        'registrations': registrations,
    })


@login_required
def qr_code_send_emails(request, event_id):
    """Send QR codes via email to all registrations"""
    event = get_object_or_404(Event, id=event_id)

    # Check permission
    if event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission.')
        return redirect('home')

    registrations = Registration.objects.filter(
        event=event,
        status__in=['pending', 'confirmed']
    ).select_related('ticket_type')

    sent_count = 0
    failed_count = 0

    for reg in registrations:
        try:
            from django.core.mail import send_mail
            from django.conf import settings

            subject = f"Your QR Code - {event.title}"
            message = f"""
Dear {reg.attendee_name},

Your registration for {event.title} has been confirmed!

Your QR Code for check-in: {reg.qr_code}
Registration Number: {reg.registration_number}

Please bring this code to the event for quick check-in.

Best regards,
{event.organizer.get_full_name() or 'Event Team'}
            """
            send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [reg.attendee_email], fail_silently=True)
            sent_count += 1
        except Exception:
            failed_count += 1

    messages.success(request, f'QR codes sent to {sent_count} attendees. {failed_count} failed.')
    return redirect('registration:qr_code_list', event_id=event.id)


# =============================================================================
# Bulk Badge Printing Views
# =============================================================================

# @event_organizer_or_admin_required
def bulk_badge_print(request, event_id):
    """Bulk print all badges for an event"""
    event = get_object_or_404(Event, id=event_id)

    registrations = Registration.objects.filter(
        event=event,
        status__in=['confirmed', 'checked_in']
    ).select_related('ticket_type', 'badge')

    badges = []
    for reg in registrations:
        # Get title and company from custom fields
        custom_fields = reg.custom_fields or {}
        title = custom_fields.get('title', '')
        company = custom_fields.get('company', '')
        
        # Create or get badge
        badge, created = Badge.objects.get_or_create(
            registration=reg,
            defaults={
                'name': reg.attendee_name,
                'title': title,
                'company': company,
                'badge_type': 'vip' if reg.ticket_type and reg.ticket_type.ticket_category == 'vip' else 'standard',
                'qr_code_data': f"BADGE:{reg.qr_code}",
            }
        )
        
        # Generate QR code
        try:
            qr_image = badge.generate_qr_code()
        except Exception:
            qr_image = None
        
        badges.append({
            'badge': badge,
            'registration': reg,
            'qr_image': qr_image,
        })

    return render(request, 'registration/bulk_badge_print.html', {
        'event': event,
        'badges': badges,
        'total_badges': len(badges),
    })


@event_organizer_or_admin_required
def bulk_badge_mark_printed(request, event_id):
    """Mark all badges as printed"""
    event = get_object_or_404(Event, id=event_id)

    registrations = Registration.objects.filter(
        event=event,
        status__in=['confirmed', 'checked_in']
    ).select_related('badge')

    printed_count = 0
    for reg in registrations:
        # Get title and company from custom fields
        custom_fields = reg.custom_fields or {}
        title = custom_fields.get('title', '')
        company = custom_fields.get('company', '')
        
        # Create or get badge
        badge, created = Badge.objects.get_or_create(
            registration=reg,
            defaults={
                'name': reg.attendee_name,
                'title': title,
                'company': company,
                'badge_type': 'vip' if reg.ticket_type and reg.ticket_type.ticket_category == 'vip' else 'standard',
                'qr_code_data': f"BADGE:{reg.qr_code}",
            }
        )
        
        # Mark as printed
        badge.mark_printed(request.user)
        printed_count += 1

    messages.success(request, f'{printed_count} badges marked as printed.')
    return redirect('registration:bulk_badge_print', event_id=event.id)


@event_organizer_or_admin_required
def bulk_badge_download_pdf(request, event_id):
    """Download all badges as PDF for printing"""
    event = get_object_or_404(Event, id=event_id)

    registrations = Registration.objects.filter(
        event=event,
        status__in=['confirmed', 'checked_in']
    ).select_related('ticket_type', 'badge')

    # Create PDF
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import inch
    from io import BytesIO
    import datetime

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Badge dimensions
    badge_width = 3.5 * inch
    badge_height = 2.25 * inch
    margin = 0.5 * inch
    
    # Calculate grid (2x3 badges per page)
    cols = 2
    rows = 3
    spacing_x = (width - 2 * margin - cols * badge_width) / (cols - 1)
    spacing_y = (height - 2 * margin - rows * badge_height) / (rows - 1)

    current_col = 0
    current_row = 0
    page_num = 1

    for reg in registrations:
        # Get title and company from custom fields
        custom_fields = reg.custom_fields or {}
        title = custom_fields.get('title', '')
        company = custom_fields.get('company', '')
        
        # Create or get badge
        badge, created = Badge.objects.get_or_create(
            registration=reg,
            defaults={
                'name': reg.attendee_name,
                'title': title,
                'company': company,
                'badge_type': 'vip' if reg.ticket_type and reg.ticket_type.ticket_category == 'vip' else 'standard',
                'qr_code_data': f"BADGE:{reg.qr_code}",
            }
        )

        # Calculate position
        x = margin + current_col * (badge_width + spacing_x)
        y = height - margin - (current_row + 1) * badge_height - current_row * spacing_y

        # Draw badge border
        p.setStrokeColorRGB(0.2, 0.2, 0.2)
        p.setFillColorRGB(1, 1, 1)
        p.rect(x, y, badge_width, badge_height, fill=1, stroke=1)

        # Badge header with color
        if badge.badge_type == 'vip':
            p.setFillColorRGB(0.8, 0.2, 0.2)  # Red for VIP
        else:
            p.setFillColorRGB(0.2, 0.4, 0.8)  # Blue for standard
        
        p.rect(x, y + badge_height - 0.5 * inch, badge_width, 0.5 * inch, fill=1, stroke=1)

        # Event name
        p.setFillColorRGB(1, 1, 1)
        p.setFont("Helvetica-Bold", 12)
        p.drawCentredText(x + badge_width/2, y + badge_height - 0.25 * inch, event.title[:30])

        # Attendee name
        p.setFillColorRGB(0, 0, 0)
        p.setFont("Helvetica-Bold", 16)
        p.drawCentredText(x + badge_width/2, y + badge_height - 0.8 * inch, badge.name[:25])

        # Title
        if badge.title:
            p.setFont("Helvetica", 10)
            p.drawCentredText(x + badge_width/2, y + badge_height - 1.1 * inch, badge.title[:30])

        # Company
        if badge.company:
            p.setFont("Helvetica", 9)
            p.drawCentredText(x + badge_width/2, y + badge_height - 1.3 * inch, badge.company[:25])

        # QR Code placeholder (small square)
        qr_size = 0.6 * inch
        qr_x = x + badge_width - qr_size - 0.1 * inch
        qr_y = y + 0.1 * inch
        p.setStrokeColorRGB(0, 0, 0)
        p.setFillColorRGB(0, 0, 0)
        p.rect(qr_x, qr_y, qr_size, qr_size, fill=0, stroke=1)

        # QR Code text
        p.setFont("Helvetica", 6)
        p.drawCentredText(qr_x + qr_size/2, qr_y - 0.05 * inch, "QR Code")

        # Registration number
        p.setFont("Helvetica", 8)
        p.drawCentredText(x + badge_width/2, y + 0.05 * inch, f"Reg: {reg.registration_number}")

        # Move to next position
        current_col += 1
        if current_col >= cols:
            current_col = 0
            current_row += 1
            if current_row >= rows:
                # New page
                p.showPage()
                page_num += 1
                current_row = 0
                # Add page number
                p.setFont("Helvetica", 8)
                p.setFillColorRGB(0.5, 0.5, 0.5)
                p.drawString(width - 1 * inch, 0.5 * inch, f"Page {page_num}")

    # Finalize PDF
    p.save()

    # Mark badges as printed
    for reg in registrations:
        badge, created = Badge.objects.get_or_create(registration=reg)
        badge.mark_printed(request.user)

    # Prepare response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="badges_{event.title}_{datetime.datetime.now().strftime("%Y%m%d")}.pdf"'
    
    return response


# =============================================================================
# Check-in Analytics Views
# =============================================================================

@login_required
def checkin_analytics(request, event_id):
    """Check-in analytics dashboard"""
    event = get_object_or_404(Event, id=event_id)

    # Check permission
    if event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission.')
        return redirect('home')

    # Get registrations
    registrations = Registration.objects.filter(event=event)

    # Stats
    total = registrations.count()
    confirmed = registrations.filter(status='confirmed').count()
    checked_in = registrations.filter(status='checked_in').count()
    pending = registrations.filter(status='pending').count()
    cancelled = registrations.filter(status='cancelled').count()

    # Check-in rate
    check_in_rate = (checked_in / confirmed * 100) if confirmed > 0 else 0

    # Check-ins by hour
    from django.db.models.functions import ExtractHour
    checkins_by_hour = CheckIn.objects.filter(
        registration__event=event
    ).annotate(
        hour=ExtractHour('check_in_time')
    ).values('hour').annotate(count=models.Count('id'))

    # Check-ins by ticket type
    checkins_by_ticket = registrations.filter(
        status='checked_in'
    ).values('ticket_type__name').annotate(count=models.Count('id'))

    # Recent check-ins
    recent_checkins = CheckIn.objects.filter(
        registration__event=event
    ).select_related('registration').order_by('-check_in_time')[:20]

    return render(request, 'registration/checkin_analytics.html', {
        'event': event,
        'total': total,
        'confirmed': confirmed,
        'checked_in': checked_in,
        'pending': pending,
        'cancelled': cancelled,
        'check_in_rate': check_in_rate,
        'checkins_by_hour': list(checkins_by_hour),
        'checkins_by_ticket': list(checkins_by_ticket),
        'recent_checkins': recent_checkins,
    })


@login_required
def checkin_history(request, event_id):
    """View complete check-in history"""
    event = get_object_or_404(Event, id=event_id)

    # Check permission
    if event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission.')
        return redirect('home')

    checkins = CheckIn.objects.filter(
        registration__event=event
    ).select_related('registration', 'checked_in_by').order_by('-check_in_time')

    # Filter by date
    date_filter = request.GET.get('date')
    if date_filter:
        checkins = checkins.filter(check_in_time__date=date_filter)

    return render(request, 'registration/checkin_history.html', {
        'event': event,
        'checkins': checkins,
        'date_filter': date_filter,
    })


@login_required
def checkin_undo(request, registration_id):
    """Undo check-in for a registration"""
    registration = get_object_or_404(Registration, id=registration_id)

    # Check permission
    if registration.event.organizer != request.user and not request.user.is_staff:
        messages.error(request, 'You do not have permission.')
        return redirect('home')

    if registration.status == 'checked_in':
        registration.status = 'confirmed'
        registration.checked_in_at = None
        registration.save()

        # Delete check-in log
        CheckIn.objects.filter(registration=registration).delete()

        messages.success(request, f'Check-in undone for {registration.attendee_name}')
    else:
        messages.warning(request, 'This registration is not checked in')

    return redirect('registration:manual_check_in', event_id=registration.event.id)
