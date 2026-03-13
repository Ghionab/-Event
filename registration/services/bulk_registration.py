"""
Enterprise Bulk Registration Services
"""
import os
import csv
import re
from io import StringIO
from decimal import Decimal
from typing import Dict, List, Tuple, Any
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook
from ..models import (
    BulkRegistrationUpload, BulkRegistrationRow, Registration, 
    RegistrationStatus, TicketType, RegistrationField, TicketAnswer,
    BulkImportAuditLog
)


class BulkRegistrationService:
    """Enterprise service for handling bulk registration operations"""
    
    def __init__(self, bulk_upload: BulkRegistrationUpload):
        self.bulk_upload = bulk_upload
        self.event = bulk_upload.event
    
    def parse_file(self) -> Tuple[List[str], List[Dict[str, Any]]]:
        """
        Parse uploaded file and extract columns and data
        Returns: (detected_columns, rows_data)
        """
        file_path = self.bulk_upload.file.path
        file_ext = os.path.splitext(self.bulk_upload.original_filename)[1].lower()
        
        try:
            if file_ext in ['.xlsx', '.xls']:
                return self._parse_excel(file_path)
            elif file_ext == '.csv':
                return self._parse_csv(file_path)
            else:
                raise ValidationError(f'Unsupported file format: {file_ext}')
        except Exception as e:
            self._log_audit('error', {'error': str(e)}, f'Failed to parse file: {str(e)}')
            raise ValidationError(f'Error parsing file: {str(e)}')
    
    def _parse_excel(self, file_path: str) -> Tuple[List[str], List[Dict[str, Any]]]:
        """Parse Excel file"""
        wb = load_workbook(filename=file_path, data_only=True)
        ws = wb.active
        
        # Get all rows as list
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise ValidationError('File is empty')
        
        # Check row limit
        if len(all_rows) > 10000:
            raise ValidationError(f'File contains {len(all_rows)} rows. Maximum allowed is 10,000 rows.')
        
        # Extract headers and data
        skip_header = self.bulk_upload.import_options.get('skip_header', True)
        
        if skip_header and len(all_rows) > 0:
            header_row = [str(cell).strip() if cell is not None else f'Column_{i+1}' 
                         for i, cell in enumerate(all_rows[0])]
            data_rows = all_rows[1:]
        else:
            # Generate column names
            max_cols = max(len(row) for row in all_rows) if all_rows else 0
            header_row = [f'Column_{i+1}' for i in range(max_cols)]
            data_rows = all_rows
        
        # Convert to dictionaries
        rows_data = []
        for row_num, row in enumerate(data_rows, 1):
            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(header_row):
                    row_dict[header_row[i]] = str(cell).strip() if cell is not None else ''
            
            # Skip empty rows
            if any(row_dict.values()):
                rows_data.append(row_dict)
        
        return header_row, rows_data
    
    def _parse_csv(self, file_path: str) -> Tuple[List[str], List[Dict[str, Any]]]:
        """Parse CSV file"""
        rows_data = []
        
        # Try different encodings
        encodings = ['utf-8', 'utf-8-sig', 'latin1', 'cp1252']
        
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding, newline='') as f:
                    # Detect delimiter
                    sample = f.read(1024)
                    f.seek(0)
                    
                    sniffer = csv.Sniffer()
                    delimiter = sniffer.sniff(sample).delimiter
                    
                    reader = csv.reader(f, delimiter=delimiter)
                    all_rows = list(reader)
                    break
            except (UnicodeDecodeError, csv.Error):
                continue
        else:
            raise ValidationError('Unable to read CSV file. Please check file encoding.')
        
        if not all_rows:
            raise ValidationError('CSV file is empty')
        
        # Check row limit
        if len(all_rows) > 10000:
            raise ValidationError(f'File contains {len(all_rows)} rows. Maximum allowed is 10,000 rows.')
        
        # Extract headers and data
        skip_header = self.bulk_upload.import_options.get('skip_header', True)
        
        if skip_header and len(all_rows) > 0:
            header_row = [str(cell).strip() for cell in all_rows[0]]
            data_rows = all_rows[1:]
        else:
            # Generate column names
            max_cols = max(len(row) for row in all_rows) if all_rows else 0
            header_row = [f'Column_{i+1}' for i in range(max_cols)]
            data_rows = all_rows
        
        # Convert to dictionaries
        for row_num, row in enumerate(data_rows, 1):
            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(header_row):
                    row_dict[header_row[i]] = str(cell).strip() if cell else ''
            
            # Skip empty rows
            if any(row_dict.values()):
                rows_data.append(row_dict)
        
        return header_row, rows_data
    
    def validate_data(self) -> Dict[str, int]:
        """
        Validate all rows and classify them
        Returns: {'valid': count, 'warning': count, 'error': count}
        """
        summary = {'valid': 0, 'warning': 0, 'error': 0}
        
        # Get existing registrations for duplicate detection
        existing_emails = set(
            Registration.objects.filter(event=self.event)
            .values_list('attendee_email', flat=True)
        )
        
        # Get ticket types for validation
        ticket_types = {tt.name.lower(): tt for tt in self.event.ticket_types.filter(is_active=True)}
        
        # Get custom fields for validation
        custom_fields = {f.field_name: f for f in self.event.registration_fields.filter(is_active=True)}
        
        rows = self.bulk_upload.rows.all().order_by('row_number')
        
        for row in rows:
            validation_errors = []
            validation_warnings = []
            
            # Apply column mapping to get mapped data
            mapped_data = self._apply_column_mapping(row.row_data)
            row.mapped_data = mapped_data
            
            # Validate required fields
            name = mapped_data.get('attendee_name', '').strip()
            email = mapped_data.get('attendee_email', '').strip()
            
            if not name:
                validation_errors.append('Attendee name is required')
            
            if not email:
                validation_errors.append('Email address is required')
            else:
                # Validate email format
                try:
                    validate_email(email)
                except ValidationError:
                    validation_errors.append('Invalid email format')
                
                # Check for duplicates
                if email.lower() in existing_emails:
                    row.is_duplicate = True
                    existing_reg = Registration.objects.filter(
                        event=self.event, 
                        attendee_email__iexact=email
                    ).first()
                    row.duplicate_of = existing_reg
                    
                    duplicate_handling = self.bulk_upload.import_options.get('duplicate_handling', 'skip')
                    if duplicate_handling == 'skip':
                        validation_warnings.append(f'Duplicate email - will be skipped')
                    elif duplicate_handling == 'update':
                        validation_warnings.append(f'Duplicate email - will update existing registration')
                    else:  # allow
                        validation_warnings.append(f'Duplicate email - will create new registration')
            
            # Validate phone number format (if provided)
            phone = mapped_data.get('attendee_phone', '').strip()
            if phone and not self._is_valid_phone(phone):
                validation_warnings.append('Phone number format may be invalid')
            
            # Validate ticket type (if mapped)
            ticket_type_name = mapped_data.get('ticket_type', '').strip()
            if ticket_type_name:
                if ticket_type_name.lower() not in ticket_types:
                    validation_errors.append(f'Ticket type "{ticket_type_name}" not found')
                else:
                    ticket_type = ticket_types[ticket_type_name.lower()]
                    row.assigned_ticket_type = ticket_type
                    
                    # Check availability
                    if ticket_type.is_sold_out:
                        validation_errors.append(f'Ticket type "{ticket_type_name}" is sold out')
                    elif not ticket_type.can_purchase():
                        validation_errors.append(f'Ticket type "{ticket_type_name}" is not available for purchase')
            
            # Validate custom fields
            for field_key, value in mapped_data.items():
                if field_key.startswith('custom_'):
                    field_id = field_key.replace('custom_', '')
                    if field_id in custom_fields:
                        field = custom_fields[field_id]
                        if field.required and not value.strip():
                            validation_errors.append(f'{field.label} is required')
            
            # Set validation status
            if validation_errors:
                row.validation_status = 'error'
                summary['error'] += 1
            elif validation_warnings:
                row.validation_status = 'warning'
                summary['warning'] += 1
            else:
                row.validation_status = 'valid'
                summary['valid'] += 1
            
            row.validation_errors = validation_errors
            row.validation_warnings = validation_warnings
            row.save()
        
        # Update bulk upload summary
        self.bulk_upload.validation_summary = summary
        self.bulk_upload.valid_rows = summary['valid']
        self.bulk_upload.warning_rows = summary['warning']
        self.bulk_upload.error_rows = summary['error']
        self.bulk_upload.current_step = 'validated'
        self.bulk_upload.save()
        
        self._log_audit('validation', summary, f'Validated {sum(summary.values())} rows')
        
        return summary
    
    def _apply_column_mapping(self, row_data: Dict[str, Any]) -> Dict[str, Any]:
        """Apply column mapping to convert file columns to system fields"""
        mapped_data = {}
        column_mapping = self.bulk_upload.column_mapping
        
        for file_column, system_field in column_mapping.items():
            if system_field and system_field != 'ignore' and file_column in row_data:
                mapped_data[system_field] = row_data[file_column]
        
        return mapped_data
    
    def _is_valid_phone(self, phone: str) -> bool:
        """Basic phone number validation"""
        # Remove common separators
        cleaned = re.sub(r'[\s\-\(\)\+\.]', '', phone)
        # Check if it's mostly digits and reasonable length
        return cleaned.isdigit() and 7 <= len(cleaned) <= 15
    
    @transaction.atomic
    def execute_import(self) -> Dict[str, int]:
        """
        Execute the actual import process
        Returns: {'success': count, 'skipped': count, 'failed': count}
        """
        self.bulk_upload.status = 'processing'
        self.bulk_upload.current_step = 'processing'
        self.bulk_upload.save()
        
        results = {'success': 0, 'skipped': 0, 'failed': 0}
        
        # Get import options
        options = self.bulk_upload.import_options
        duplicate_handling = options.get('duplicate_handling', 'skip')
        send_emails = options.get('send_emails', True)
        default_ticket_type = self.bulk_upload.default_ticket_type
        
        # Process valid and warning rows
        processable_rows = self.bulk_upload.rows.filter(
            validation_status__in=['valid', 'warning']
        ).order_by('row_number')
        
        for row in processable_rows:
            try:
                # Handle duplicates
                if row.is_duplicate:
                    if duplicate_handling == 'skip':
                        row.status = 'skipped'
                        row.processed_at = timezone.now()
                        row.save()
                        results['skipped'] += 1
                        continue
                    elif duplicate_handling == 'update':
                        # Update existing registration
                        self._update_existing_registration(row)
                        results['success'] += 1
                        continue
                    # If 'allow', continue to create new registration
                
                # Create new registration
                registration = self._create_registration(row, default_ticket_type)
                
                # Send email if requested
                if send_emails and registration:
                    self._send_confirmation_email(registration)
                
                row.registration = registration
                row.status = 'success'
                row.processed_at = timezone.now()
                row.save()
                results['success'] += 1
                
            except Exception as e:
                row.status = 'failed'
                row.error_message = str(e)
                row.processed_at = timezone.now()
                row.save()
                results['failed'] += 1
        
        # Update bulk upload results
        self.bulk_upload.success_count = results['success']
        self.bulk_upload.skipped_count = results['skipped']
        self.bulk_upload.error_count = results['failed']
        self.bulk_upload.status = 'completed'
        self.bulk_upload.current_step = 'completed'
        self.bulk_upload.completed_at = timezone.now()
        self.bulk_upload.save()
        
        self._log_audit('completion', results, 
                       f'Import completed: {results["success"]} success, {results["skipped"]} skipped, {results["failed"]} failed')
        
        return results
    
    def _create_registration(self, row: BulkRegistrationRow, default_ticket_type: TicketType) -> Registration:
        """Create a new registration from row data"""
        mapped_data = row.mapped_data
        
        # Determine ticket type
        ticket_type = row.assigned_ticket_type or default_ticket_type
        if not ticket_type:
            raise ValidationError('No ticket type available')
        
        # Create registration
        registration = Registration.objects.create(
            event=self.event,
            attendee_name=mapped_data.get('attendee_name', ''),
            attendee_email=mapped_data.get('attendee_email', ''),
            attendee_phone=mapped_data.get('attendee_phone', ''),
            ticket_type=ticket_type,
            total_amount=ticket_type.price,
            status=RegistrationStatus.CONFIRMED,
            custom_fields={
                'company': mapped_data.get('company', ''),
                'job_title': mapped_data.get('job_title', ''),
                'bulk_import': True,
                'imported_by': self.bulk_upload.uploaded_by.id if self.bulk_upload.uploaded_by else None,
            }
        )
        
        # Update ticket count
        ticket_type.quantity_sold += 1
        ticket_type.save()
        
        # Create custom field answers
        self._create_custom_answers(registration, mapped_data)
        
        return registration
    
    def _update_existing_registration(self, row: BulkRegistrationRow):
        """Update existing registration with new data"""
        if not row.duplicate_of:
            raise ValidationError('No existing registration found for update')
        
        registration = row.duplicate_of
        mapped_data = row.mapped_data
        
        # Update fields
        registration.attendee_name = mapped_data.get('attendee_name', registration.attendee_name)
        registration.attendee_phone = mapped_data.get('attendee_phone', registration.attendee_phone)
        
        # Update custom fields
        if not registration.custom_fields:
            registration.custom_fields = {}
        
        registration.custom_fields.update({
            'company': mapped_data.get('company', registration.custom_fields.get('company', '')),
            'job_title': mapped_data.get('job_title', registration.custom_fields.get('job_title', '')),
            'bulk_updated': True,
            'updated_by': self.bulk_upload.uploaded_by.id if self.bulk_upload.uploaded_by else None,
        })
        
        registration.save()
        
        # Update custom field answers
        self._create_custom_answers(registration, mapped_data)
        
        row.registration = registration
        row.status = 'success'
        row.processed_at = timezone.now()
        row.save()
    
    def _create_custom_answers(self, registration: Registration, mapped_data: Dict[str, Any]):
        """Create answers for custom registration fields"""
        for field_key, value in mapped_data.items():
            if field_key.startswith('custom_') and value.strip():
                field_id = field_key.replace('custom_', '')
                try:
                    field = RegistrationField.objects.get(id=field_id, event=self.event)
                    # Note: This assumes we have a Ticket model with answers relationship
                    # If not, we can store in Registration.custom_fields instead
                    pass  # Implementation depends on your ticket/answer model structure
                except RegistrationField.DoesNotExist:
                    pass
    
    def _send_confirmation_email(self, registration: Registration):
        """Send confirmation email to attendee"""
        try:
            from django.core.mail import send_mail
            from django.conf import settings
            
            subject = f"Registration Confirmed: {self.event.title}"
            message = f"""
Dear {registration.attendee_name},

Your registration for {self.event.title} has been confirmed!

Event Details:
- Date: {self.event.start_date.strftime('%B %d, %Y') if hasattr(self.event, 'start_date') else 'TBA'}
- Venue: {getattr(self.event, 'venue_name', 'TBA')}
- Ticket: {registration.ticket_type.name}

Registration Number: {registration.registration_number}

Best regards,
{self.bulk_upload.uploaded_by.get_full_name() if self.bulk_upload.uploaded_by else 'Event Team'}
            """
            
            send_mail(
                subject, 
                message, 
                settings.DEFAULT_FROM_EMAIL, 
                [registration.attendee_email], 
                fail_silently=True
            )
        except Exception:
            pass  # Don't fail import if email fails
    
    def _log_audit(self, action_type: str, action_details: Dict[str, Any], summary: str):
        """Log audit trail"""
        BulkImportAuditLog.objects.create(
            bulk_upload=self.bulk_upload,
            action_type=action_type,
            action_details=action_details,
            performed_by=self.bulk_upload.uploaded_by,
            result_summary=summary
        )
    
    def generate_error_report(self) -> str:
        """Generate CSV error report for download"""
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            'Row Number', 'Status', 'Attendee Name', 'Email', 
            'Validation Errors', 'Validation Warnings', 'Processing Error'
        ])
        
        # Write error and warning rows
        problem_rows = self.bulk_upload.rows.filter(
            validation_status__in=['error', 'warning']
        ).order_by('row_number')
        
        for row in problem_rows:
            mapped_data = row.mapped_data or {}
            writer.writerow([
                row.row_number,
                row.get_validation_status_display(),
                mapped_data.get('attendee_name', ''),
                mapped_data.get('attendee_email', ''),
                '; '.join(row.validation_errors or []),
                '; '.join(row.validation_warnings or []),
                row.error_message or ''
            ])
        
        return output.getvalue()