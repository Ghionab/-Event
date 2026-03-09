from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse, FileResponse
from django.utils import timezone
from django.db.models import Sum, Count, Avg
from django.conf import settings
from .decorators import organizer_required
from .models import (
    BusinessSponsor, SponsorMaterial, SponsorLead, EventAnalytics, SessionAnalytics,
    Expense, Budget, Invoice, Quote, Report, ReportExport
)
from .forms import SponsorForm, ExpenseForm, BudgetForm, InvoiceForm, QuoteForm, ReportForm
from events.models import Event
from registration.models import Registration, RegistrationStatus


# ============ Sponsor Management ============

@organizer_required
def sponsor_list(request):
    """List all sponsors"""
    event_id = request.GET.get('event')
    if event_id:
        sponsors = BusinessSponsor.objects.filter(event_id=event_id)
    else:
        sponsors = BusinessSponsor.objects.all()
    
    return render(request, 'business/sponsor_list.html', {'sponsors': sponsors})


@organizer_required
def sponsor_create(request):
    """Create a new sponsor"""
    if request.method == 'POST':
        form = SponsorForm(request.POST, request.FILES)
        if form.is_valid():
            sponsor = form.save()
            messages.success(request, f'Sponsor {sponsor.company_name} created successfully!')
            return redirect('business:sponsor_detail', sponsor_id=sponsor.id)
    else:
        form = SponsorForm()
        if 'event_id' in request.GET:
            form.fields['event'].initial = request.GET['event_id']
    
    return render(request, 'business/sponsor_form.html', {'form': form})


@organizer_required
def sponsor_detail(request, sponsor_id):
    """View sponsor details"""
    sponsor = get_object_or_404(BusinessSponsor, id=sponsor_id)
    materials = sponsor.materials.all()
    leads = sponsor.leads.all()
    
    lead_stats = {
        'total': leads.count(),
        'new': leads.filter(follow_up_status='new').count(),
        'contacted': leads.filter(follow_up_status='contacted').count(),
        'qualified': leads.filter(follow_up_status='qualified').count(),
    }
    
    return render(request, 'business/sponsor_detail.html', {
        'sponsor': sponsor,
        'materials': materials,
        'leads': leads,
        'lead_stats': lead_stats,
    })


@organizer_required
def sponsor_edit(request, sponsor_id):
    """Edit a sponsor"""
    sponsor = get_object_or_404(BusinessSponsor, id=sponsor_id)
    
    if request.method == 'POST':
        form = SponsorForm(request.POST, request.FILES, instance=sponsor)
        if form.is_valid():
            form.save()
            messages.success(request, 'Sponsor updated successfully!')
            return redirect('business:sponsor_detail', sponsor_id=sponsor.id)
    else:
        form = SponsorForm(instance=sponsor)
    
    return render(request, 'business/sponsor_form.html', {'form': form, 'sponsor': sponsor})


@organizer_required
def sponsor_delete(request, sponsor_id):
    """Delete a sponsor"""
    sponsor = get_object_or_404(BusinessSponsor, id=sponsor_id)
    
    if request.method == 'POST':
        sponsor_name = sponsor.company_name
        sponsor.delete()
        messages.success(request, f'Sponsor {sponsor_name} deleted successfully!')
        return redirect('business:sponsor_list')
    
    return render(request, 'business/confirm_delete.html', {'object': sponsor})


@organizer_required
def sponsor_materials(request, sponsor_id):
    """Manage sponsor materials"""
    sponsor = get_object_or_404(BusinessSponsor, id=sponsor_id)
    
    if request.method == 'POST':
        title = request.POST.get('title')
        file = request.FILES.get('file')
        if title and file:
            SponsorMaterial.objects.create(sponsor=sponsor, title=title, file=file)
            messages.success(request, 'Material uploaded successfully!')
    
    materials = sponsor.materials.all()
    return render(request, 'business/sponsor_materials.html', {'sponsor': sponsor, 'materials': materials})


@organizer_required
def sponsor_leads(request, sponsor_id):
    """View and manage sponsor leads"""
    sponsor = get_object_or_404(BusinessSponsor, id=sponsor_id)
    leads = sponsor.leads.all()
    
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        phone = request.POST.get('phone', '')
        company = request.POST.get('company', '')
        
        if name and email:
            SponsorLead.objects.create(
                sponsor=sponsor,
                event=sponsor.event,
                name=name,
                email=email,
                phone=phone,
                company=company
            )
            messages.success(request, 'Lead added successfully!')
    
    return render(request, 'business/sponsor_leads.html', {'sponsor': sponsor, 'leads': leads})


@organizer_required
def import_leads(request, sponsor_id):
    """Import leads from file"""
    sponsor = get_object_or_404(BusinessSponsor, id=sponsor_id)
    
    if request.method == 'POST':
        file = request.FILES.get('file')
        if file:
            content = file.read().decode('utf-8')
            lines = content.split('\n')
            imported = 0
            for line in lines[1:]:
                if line.strip():
                    parts = line.split(',')
                    if len(parts) >= 2:
                        SponsorLead.objects.create(
                            sponsor=sponsor,
                            event=sponsor.event,
                            name=parts[0].strip(),
                            email=parts[1].strip(),
                            company=parts[2].strip() if len(parts) > 2 else ''
                        )
                        imported += 1
            messages.success(request, f'Imported {imported} leads!')

    return redirect('business:sponsor_leads', sponsor_id=sponsor.id)


# ============ Analytics ============

@organizer_required
def analytics_overview(request):
    """Overview of all analytics"""
    events = Event.objects.filter(organizer=request.user)
    
    total_registrations = 0
    total_revenue = 0
    total_checked_in = 0
    
    for event in events:
        analytics, created = EventAnalytics.objects.get_or_create(event=event)
        total_registrations += analytics.total_registrations
        total_revenue += analytics.total_revenue
        total_checked_in += analytics.checked_in_count
    
    context = {
        'events': events,
        'total_registrations': total_registrations,
        'total_revenue': total_revenue,
        'total_checked_in': total_checked_in,
    }
    return render(request, 'business/analytics_overview.html', context)


@organizer_required
def event_analytics(request, event_id):
    """Detailed analytics for an event"""
    event = get_object_or_404(Event, id=event_id)
    analytics, created = EventAnalytics.objects.get_or_create(event=event)
    analytics.update_stats()
    
    reg_by_status = Registration.objects.filter(event=event).values('status').annotate(count=Count('id'))
    revenue_by_ticket = Registration.objects.filter(
        event=event, status='confirmed'
    ).values('ticket_type__name').annotate(total=Sum('total_amount'))
    
    from django.db.models.functions import TruncDate
    daily_regs = Registration.objects.filter(
        event=event
    ).annotate(date=TruncDate('created_at')).values('date').annotate(count=Count('id')).order_by('-date')[:30]
    
    context = {
        'event': event,
        'analytics': analytics,
        'reg_by_status': reg_by_status,
        'revenue_by_ticket': revenue_by_ticket,
        'daily_regs': daily_regs,
    }
    return render(request, 'business/event_analytics.html', context)


@organizer_required
def export_analytics(request, event_id):
    """Export analytics data"""
    event = get_object_or_404(Event, id=event_id)
    analytics = get_object_or_404(EventAnalytics, event=event)
    
    csv_data = f"""Event Analytics Report
Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}

Registration Statistics
Total Registrations: {analytics.total_registrations}
Confirmed: {analytics.confirmed_registrations}
Cancelled: {analytics.cancelled_registrations}
Checked In: {analytics.checked_in_count}

Revenue Statistics
Total Revenue: ${analytics.total_revenue}
Total Refunds: ${analytics.total_refunds}
Net Revenue: ${analytics.net_revenue}
"""
    
    response = FileResponse(csv_data.encode(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="analytics_{event.slug}.csv"'
    return response


# ============ Expenses ============

@organizer_required
def expense_list(request):
    """List all expenses"""
    event_id = request.GET.get('event')
    if event_id:
        expenses = Expense.objects.filter(event_id=event_id)
    else:
        expenses = Expense.objects.all()
    
    total = expenses.aggregate(total=Sum('amount'))['total'] or 0
    by_category = expenses.values('category').annotate(total=Sum('amount'))
    
    return render(request, 'business/expense_list.html', {
        'expenses': expenses,
        'total': total,
        'by_category': by_category,
    })


@organizer_required
def expense_create(request):
    """Create a new expense"""
    if request.method == 'POST':
        form = ExpenseForm(request.POST, request.FILES)
        if form.is_valid():
            expense = form.save()
            messages.success(request, 'Expense added successfully!')
            return redirect('business:expense_list')
    else:
        form = ExpenseForm()
        if 'event_id' in request.GET:
            form.fields['event'].initial = request.GET['event_id']
    
    return render(request, 'business/expense_form.html', {'form': form})


@organizer_required
def expense_edit(request, expense_id):
    """Edit an expense"""
    expense = get_object_or_404(Expense, id=expense_id)
    
    if request.method == 'POST':
        form = ExpenseForm(request.POST, request.FILES, instance=expense)
        if form.is_valid():
            form.save()
            messages.success(request, 'Expense updated successfully!')
            return redirect('business:expense_list')
    else:
        form = ExpenseForm(instance=expense)
    
    return render(request, 'business/expense_form.html', {'form': form, 'expense': expense})


@organizer_required
def expense_delete(request, expense_id):
    """Delete an expense"""
    expense = get_object_or_404(Expense, id=expense_id)
    
    if request.method == 'POST':
        expense.delete()
        messages.success(request, 'Expense deleted successfully!')
        return redirect('business:expense_list')
    
    return render(request, 'business/confirm_delete.html', {'object': expense})


# ============ Budget ============

@organizer_required
def budget_list(request):
    """List budgets"""
    event_id = request.GET.get('event')
    if event_id:
        budgets = Budget.objects.filter(event_id=event_id)
    else:
        budgets = Budget.objects.all()
    
    total_planned = sum(b.planned_amount for b in budgets)
    total_actual = sum(b.actual_amount for b in budgets)
    
    return render(request, 'business/budget_list.html', {
        'budgets': budgets,
        'total_planned': total_planned,
        'total_actual': total_actual,
    })


@organizer_required
def budget_create(request):
    """Create a budget"""
    if request.method == 'POST':
        form = BudgetForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Budget created successfully!')
            return redirect('business:budget_list')
    else:
        form = BudgetForm()
        if 'event_id' in request.GET:
            form.fields['event'].initial = request.GET['event_id']
    
    return render(request, 'business/budget_form.html', {'form': form})


@organizer_required
def budget_edit(request, budget_id):
    """Edit a budget"""
    budget = get_object_or_404(Budget, id=budget_id)
    
    if request.method == 'POST':
        form = BudgetForm(request.POST, instance=budget)
        if form.is_valid():
            form.save()
            messages.success(request, 'Budget updated successfully!')
            return redirect('business:budget_list')
    else:
        form = BudgetForm(instance=budget)
    
    return render(request, 'business/budget_form.html', {'form': form, 'budget': budget})


# ============ Invoices ============

@organizer_required
def invoice_list(request):
    """List invoices"""
    invoices = Invoice.objects.all()
    
    total_outstanding = invoices.filter(status__in=['sent', 'viewed', 'overdue']).aggregate(
        total=Sum('total_amount')
    )['total'] or 0
    total_paid = invoices.filter(status='paid').aggregate(total=Sum('total_amount'))['total'] or 0
    
    return render(request, 'business/invoice_list.html', {
        'invoices': invoices,
        'total_outstanding': total_outstanding,
        'total_paid': total_paid,
    })


@organizer_required
def invoice_create(request):
    """Create an invoice"""
    if request.method == 'POST':
        form = InvoiceForm(request.POST)
        if form.is_valid():
            invoice = form.save()
            messages.success(request, f'Invoice {invoice.invoice_number} created!')
            return redirect('business:invoice_list')
    else:
        form = InvoiceForm()
        if 'event_id' in request.GET:
            form.fields['event'].initial = request.GET['event_id']
    
    return render(request, 'business/invoice_form.html', {'form': form})


@organizer_required
def invoice_detail(request, invoice_id):
    """View invoice details"""
    invoice = get_object_or_404(Invoice, id=invoice_id)
    return render(request, 'business/invoice_detail.html', {'invoice': invoice})


@organizer_required
def invoice_edit(request, invoice_id):
    """Edit an invoice"""
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    if request.method == 'POST':
        form = InvoiceForm(request.POST, instance=invoice)
        if form.is_valid():
            form.save()
            messages.success(request, 'Invoice updated!')
            return redirect('business:invoice_list')
    else:
        form = InvoiceForm(instance=invoice)
    
    return render(request, 'business/invoice_form.html', {'form': form, 'invoice': invoice})


@organizer_required
def invoice_send(request, invoice_id):
    """Send invoice to client"""
    invoice = get_object_or_404(Invoice, id=invoice_id)
    invoice.status = 'sent'
    invoice.save()
    messages.success(request, f'Invoice {invoice.invoice_number} sent to {invoice.client_email}')
    return redirect('business:invoice_list')


@organizer_required
def invoice_mark_paid(request, invoice_id):
    """Mark invoice as paid"""
    invoice = get_object_or_404(Invoice, id=invoice_id)
    invoice.status = 'paid'
    invoice.paid_at = timezone.now()
    invoice.save()
    messages.success(request, f'Invoice {invoice.invoice_number} marked as paid')
    return redirect('business:invoice_list')


# ============ Quotes ============

@organizer_required
def quote_list(request):
    """List quotes"""
    quotes = Quote.objects.all()
    return render(request, 'business/quote_list.html', {'quotes': quotes})


@organizer_required
def quote_create(request):
    """Create a quote"""
    if request.method == 'POST':
        form = QuoteForm(request.POST)
        if form.is_valid():
            quote = form.save()
            messages.success(request, f'Quote {quote.quote_number} created!')
            return redirect('business:quote_list')
    else:
        form = QuoteForm()
    
    return render(request, 'business/quote_form.html', {'form': form})


@organizer_required
def quote_detail(request, quote_id):
    """View quote details"""
    quote = get_object_or_404(Quote, id=quote_id)
    return render(request, 'business/quote_detail.html', {'quote': quote})


@organizer_required
def quote_edit(request, quote_id):
    """Edit a quote"""
    quote = get_object_or_404(Quote, id=quote_id)
    
    if request.method == 'POST':
        form = QuoteForm(request.POST, instance=quote)
        if form.is_valid():
            form.save()
            messages.success(request, 'Quote updated!')
            return redirect('business:quote_list')
    else:
        form = QuoteForm(instance=quote)
    
    return render(request, 'business/quote_form.html', {'form': form, 'quote': quote})


@organizer_required
def quote_convert(request, quote_id):
    """Convert quote to invoice"""
    quote = get_object_or_404(Quote, id=quote_id)
    
    invoice = Invoice.objects.create(
        event=quote.event,
        invoice_number=f"INV-{quote.quote_number.split('-')[1]}",
        client_name=quote.client_name,
        client_email=quote.client_email,
        description=quote.notes or 'Services from quote',
        amount=quote.total_amount,
        tax_rate=quote.tax_rate,
        issue_date=timezone.now().date(),
        due_date=quote.valid_until,
        status='draft'
    )
    
    quote.status = 'accepted'
    quote.save()
    
    messages.success(request, f'Quote converted to Invoice {invoice.invoice_number}')
    return redirect('business:invoice_detail', invoice_id=invoice.id)


# ============ Reports ============

@organizer_required
def report_list(request):
    """List reports"""
    reports = Report.objects.all()
    return render(request, 'business/report_list.html', {'reports': reports})


@organizer_required
def report_create(request):
    """Create a report"""
    if request.method == 'POST':
        form = ReportForm(request.POST)
        if form.is_valid():
            report = form.save(commit=False)
            report.created_by = request.user
            report.save()
            messages.success(request, 'Report created!')
            return redirect('business:report_list')
    else:
        form = ReportForm()
        if 'event_id' in request.GET:
            form.fields['event'].initial = request.GET['event_id']
    
    return render(request, 'business/report_form.html', {'form': form})


@organizer_required
def report_detail(request, report_id):
    """View report details"""
    report = get_object_or_404(Report, id=report_id)
    exports = report.exports.all()
    return render(request, 'business/report_detail.html', {'report': report, 'exports': exports})


@organizer_required
def report_generate(request, report_id):
    """Generate report data"""
    report = get_object_or_404(Report, id=report_id)
    
    # Parse filters if they exist
    filters = {}
    if report.filters:
        try:
            import json
            filters = json.loads(report.filters) if isinstance(report.filters, str) else report.filters
        except:
            filters = report.filters if isinstance(report.filters, dict) else {}
    
    # Start with base queryset
    regs = Registration.objects.filter(event=report.event)
    
    # Apply filters
    if filters:
        # Status filter
        if 'status' in filters:
            regs = regs.filter(status=filters['status'])
        
        # Ticket type filter
        if 'ticket_type' in filters:
            regs = regs.filter(ticket_type__name__icontains=filters['ticket_type'])
        
        # Payment status filter
        if 'payment_status' in filters:
            regs = regs.filter(payment_status=filters['payment_status'])
        
        # Date range filter
        if 'date_range' in filters:
            date_range = filters['date_range']
            if 'from' in date_range:
                regs = regs.filter(created_at__gte=date_range['from'])
            if 'to' in date_range:
                regs = regs.filter(created_at__lte=date_range['to'])
    
    # Generate data based on report type
    data = {}
    
    if report.report_type == 'registration':
        data = {
            'total': regs.count(),
            'by_status': list(regs.values('status').annotate(count=Count('id'))),
            'recent': list(regs.order_by('-created_at')[:10].values(
                'registration_number', 'attendee_name', 'attendee_email', 
                'status', 'total_amount', 'created_at'
            )),
        }
    elif report.report_type == 'revenue':
        confirmed_regs = regs.filter(status='confirmed')
        data = {
            'total_revenue': float(sum(r.total_amount for r in confirmed_regs)),
            'total_registrations': confirmed_regs.count(),
            'by_ticket': list(confirmed_regs.values('ticket_type__name').annotate(
                total=Sum('total_amount'),
                count=Count('id')
            )),
            'average_ticket_price': float(confirmed_regs.aggregate(avg=Avg('total_amount'))['avg'] or 0),
        }
    elif report.report_type == 'attendance':
        data = {
            'total': regs.count(),
            'by_status': list(regs.values('status').annotate(count=Count('id'))),
            'checked_in': regs.filter(status='checked_in').count(),
            'confirmed': regs.filter(status='confirmed').count(),
            'cancelled': regs.filter(status__in=['cancelled', 'refunded']).count(),
        }
    
    # Update report metadata
    report.last_generated = timezone.now()
    report.report_data = data
    report.save()
    
    # Return JSON for AJAX or redirect for regular request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'data': data})
    else:
        messages.success(request, f'Report "{report.name}" generated successfully!')
        return redirect('business:report_detail', report_id=report.id)


@organizer_required
def report_export(request, report_id):
    """Export report in multiple formats"""
    report = get_object_or_404(Report, id=report_id)
    format_type = request.GET.get('format', report.export_format)
    
    # Generate report data
    data = []
    headers = []
    
    if report.report_type == 'registration':
        regs = Registration.objects.filter(event=report.event)
        headers = ['Registration Number', 'Name', 'Email', 'Status', 'Ticket Type', 'Amount', 'Date']
        for r in regs:
            data.append([
                r.registration_number,
                r.attendee_name,
                r.attendee_email,
                r.get_status_display(),
                r.ticket_type.name if r.ticket_type else 'N/A',
                f"${r.total_amount}",
                r.created_at.strftime('%Y-%m-%d %H:%M')
            ])
    
    elif report.report_type == 'revenue':
        regs = Registration.objects.filter(event=report.event, status='confirmed')
        headers = ['Ticket Type', 'Quantity', 'Total Revenue']
        ticket_data = {}
        for r in regs:
            ticket_name = r.ticket_type.name if r.ticket_type else 'N/A'
            if ticket_name not in ticket_data:
                ticket_data[ticket_name] = {'count': 0, 'revenue': 0}
            ticket_data[ticket_name]['count'] += 1
            ticket_data[ticket_name]['revenue'] += r.total_amount
        
        for ticket_name, stats in ticket_data.items():
            data.append([ticket_name, stats['count'], f"${stats['revenue']}"])
    
    elif report.report_type == 'attendance':
        regs = Registration.objects.filter(event=report.event)
        headers = ['Status', 'Count', 'Percentage']
        total = regs.count()
        status_counts = {}
        for r in regs:
            status = r.get_status_display()
            status_counts[status] = status_counts.get(status, 0) + 1
        
        for status, count in status_counts.items():
            percentage = (count / total * 100) if total > 0 else 0
            data.append([status, count, f"{percentage:.1f}%"])
    
    # Export based on format
    if format_type == 'csv':
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{report.name.replace(" ", "_")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(data)
        
        return response
    
    elif format_type == 'xlsx':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from django.http import HttpResponse
            from io import BytesIO
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = report.name[:31]  # Excel sheet name limit
            
            # Add headers with styling
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            # Add data
            for row_num, row_data in enumerate(data, 2):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{report.name.replace(" ", "_")}.xlsx"'
            
            return response
        
        except ImportError:
            messages.error(request, 'Excel export requires openpyxl package. Install with: pip install openpyxl')
            return redirect('business:report_detail', report_id=report.id)
    
    elif format_type == 'pdf':
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from django.http import HttpResponse
            from io import BytesIO
            
            # Create PDF
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#366092'),
                spaceAfter=30,
            )
            
            # Add title
            elements.append(Paragraph(report.name, title_style))
            elements.append(Spacer(1, 0.2*inch))
            
            # Add event info
            info_style = styles['Normal']
            elements.append(Paragraph(f"<b>Event:</b> {report.event.title}", info_style))
            elements.append(Paragraph(f"<b>Generated:</b> {timezone.now().strftime('%Y-%m-%d %H:%M')}", info_style))
            elements.append(Spacer(1, 0.3*inch))
            
            # Add table
            table_data = [headers] + data
            table = Table(table_data)
            
            # Table styling
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            elements.append(table)
            
            # Build PDF
            doc.build(elements)
            
            # Get PDF value
            pdf = buffer.getvalue()
            buffer.close()
            
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{report.name.replace(" ", "_")}.pdf"'
            response.write(pdf)
            
            return response
        
        except ImportError:
            messages.error(request, 'PDF export requires reportlab package. Install with: pip install reportlab')
            return redirect('business:report_detail', report_id=report.id)
    
    # Default to CSV if format not recognized
    messages.error(request, f'Unsupported format: {format_type}')
    return redirect('business:report_detail', report_id=report.id)


# ============ Dashboard ============

@organizer_required
def business_dashboard(request):
    """Business overview dashboard"""
    events = Event.objects.filter(organizer=request.user)
    
    total_sponsors = BusinessSponsor.objects.filter(event__in=events).count()
    total_sponsorship = BusinessSponsor.objects.filter(event__in=events).aggregate(
        total=Sum('sponsorship_amount')
    )['total'] or 0
    
    total_expenses = Expense.objects.filter(event__in=events).aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    total_invoices = Invoice.objects.filter(event__in=events).aggregate(
        total=Sum('total_amount')
    )['total'] or 0
    paid_invoices = Invoice.objects.filter(event__in=events, status='paid').aggregate(
        total=Sum('total_amount')
    )['total'] or 0
    
    recent_sponsors = BusinessSponsor.objects.filter(event__in=events).order_by('-created_at')[:5]
    recent_expenses = Expense.objects.filter(event__in=events).order_by('-created_at')[:5]
    recent_invoices = Invoice.objects.filter(event__in=events).order_by('-created_at')[:5]
    
    context = {
        'events': events,
        'total_sponsors': total_sponsors,
        'total_sponsorship': total_sponsorship,
        'total_expenses': total_expenses,
        'total_invoices': total_invoices,
        'paid_invoices': paid_invoices,
        'recent_sponsors': recent_sponsors,
        'recent_expenses': recent_expenses,
        'recent_invoices': recent_invoices,
    }
    return render(request, 'business/dashboard.html', context)
